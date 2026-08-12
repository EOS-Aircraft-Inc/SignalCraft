"""Thin wrappers around csv_to_excel / excel_to_csv for the Streamlit sidebar."""

from __future__ import annotations

import tempfile
from pathlib import Path

from csv_to_excel import rebuild_workbook
from excel_to_csv import export_workbook
from icd_paths import DEFAULT_CSV_DIR, DEFAULT_REBUILT
from icd_sheets import (
    DATABUSES_SHEET,
    SIGNALS_SHEET,
    SYSTEMS_SHEET,
)
from openpyxl import load_workbook  # pyright: ignore[reportMissingModuleSource]

REQUIRED_SHEETS = (SYSTEMS_SHEET, SIGNALS_SHEET, DATABUSES_SHEET)

__all__ = [
    "DEFAULT_CSV_DIR",
    "DEFAULT_REBUILT",
    "REQUIRED_SHEETS",
    "IncompatibleWorkbookError",
    "export_csv_to_excel",
    "import_excel_to_csv",
    "validate_icd_workbook",
]


class IncompatibleWorkbookError(ValueError):
    """Workbook is missing required ICD tabs."""


def validate_icd_workbook(workbook_path: Path) -> list[str]:
    """Ensure required catalog tabs exist; return sheet names.

    Raises ``IncompatibleWorkbookError`` when ``0_Systems``, ``1_Signals``, or
    ``10_Databuses`` is missing.
    """
    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        names = [str(title) for title in workbook.sheetnames]
    finally:
        workbook.close()
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in names]
    if missing:
        raise IncompatibleWorkbookError(
            "Incompatible file: missing required tab(s) "
            + ", ".join(missing)
            + f". An ICD workbook must include {', '.join(REQUIRED_SHEETS)}."
        )
    return names


def export_csv_to_excel(
    output_path: Path,
    *,
    csv_dir: Path = DEFAULT_CSV_DIR,
) -> Path:
    """Rebuild an Excel workbook from ``csv/`` into ``output_path``."""
    path = Path(output_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        path = path.with_suffix(".xlsx")
    rebuild_workbook(Path(csv_dir).resolve(), path.resolve())
    return path.resolve()


def import_excel_to_csv(
    workbook_path: Path | bytes,
    *,
    csv_dir: Path = DEFAULT_CSV_DIR,
    source_name: str = "upload.xlsx",
) -> Path:
    """Export an Excel workbook into ``csv/`` (overwrites the working CSV set)."""
    out = Path(csv_dir).resolve()
    if isinstance(workbook_path, (bytes, bytearray)):
        safe_name = Path(source_name).name or "upload.xlsx"
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp) / safe_name
            if tmp_path.suffix.lower() not in {".xlsx", ".xlsm"}:
                tmp_path = tmp_path.with_suffix(".xlsx")
            tmp_path.write_bytes(bytes(workbook_path))
            validate_icd_workbook(tmp_path)
            export_workbook(tmp_path.resolve(), out)
        return out

    path = Path(workbook_path).resolve()
    validate_icd_workbook(path)
    export_workbook(path, out)
    return out
