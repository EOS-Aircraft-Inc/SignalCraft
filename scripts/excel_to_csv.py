"""Export every ICD_Database.xlsx worksheet to CSV (SignalCraft).

The Excel workbook remains the source of truth. Run this script before using
text-based tools or AI on the ICD data.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook  # pyright: ignore[reportMissingModuleSource]

from icd_csv import safe_filename
from icd_paths import DEFAULT_CSV_DIR, DEFAULT_WORKBOOK, MANIFEST_NAME


def export_workbook(workbook_path: Path, output_dir: Path) -> None:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        manifest: dict[str, object] = {
            "source_workbook": workbook_path.name,
            "sheets": [],
        }
        expected_files: set[str] = set()

        for index, worksheet in enumerate(workbook.worksheets, start=1):
            filename = f"{index:02d}_{safe_filename(worksheet.title)}.csv"
            expected_files.add(filename)
            csv_path = output_dir / filename

            with csv_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
                writer = csv.writer(csv_file)
                for row in worksheet.iter_rows(values_only=True):
                    writer.writerow(
                        ["" if value is None else value for value in row]
                    )

            manifest["sheets"].append(
                {
                    "order": index,
                    "sheet_name": worksheet.title,
                    "csv_file": filename,
                    "sheet_state": worksheet.sheet_state,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                }
            )
            print(f"Exported {worksheet.title!r} -> {csv_path.name}")

        for csv_path in output_dir.glob("*.csv"):
            if csv_path.name not in expected_files:
                csv_path.unlink()
                print(f"Removed stale export {csv_path.name}")

        manifest_path = output_dir / MANIFEST_NAME
        manifest_path.write_text(
            json.dumps(manifest, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        print(f"Manifest -> {manifest_path}")
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export all ICD Excel worksheets to individual CSV files."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Source workbook (default: {DEFAULT_WORKBOOK.name})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_CSV_DIR,
        help=f"CSV output directory (default: {DEFAULT_CSV_DIR})",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    export_workbook(args.workbook.resolve(), args.output_dir.resolve())
