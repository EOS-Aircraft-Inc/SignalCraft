"""Rebuild an ICD workbook from the generated per-sheet CSV files (SignalCraft).

This creates a new workbook by default. It never overwrites ICD_Database.xlsx
unless the user explicitly supplies that path with --output.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import TypedDict

from icd_paths import DEFAULT_CSV_DIR, DEFAULT_REBUILT, DEFAULT_WORKBOOK, MANIFEST_NAME
from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import (  # pyright: ignore[reportMissingModuleSource]
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter  # pyright: ignore[reportMissingModuleSource]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class SheetManifestEntry(TypedDict):
    order: int
    sheet_name: str
    csv_file: str
    sheet_state: str


def load_manifest(input_dir: Path) -> list[SheetManifestEntry]:
    manifest_path = input_dir / MANIFEST_NAME
    if not manifest_path.is_file():
        raise FileNotFoundError(
            f"Manifest not found: {manifest_path}. Run excel_to_csv.py first."
        )

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    sheets = manifest.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError(f"No worksheets listed in {manifest_path}")
    entries: list[SheetManifestEntry] = []
    for sheet in sheets:
        if not isinstance(sheet, dict):
            raise TypeError(f"Invalid worksheet entry in {manifest_path}")
        entries.append(
            SheetManifestEntry(
                order=int(str(sheet["order"])),
                sheet_name=str(sheet["sheet_name"]),
                csv_file=str(sheet["csv_file"]),
                sheet_state=str(sheet.get("sheet_state", "visible")),
            )
        )
    return sorted(entries, key=lambda sheet: sheet["order"])


def format_worksheet(worksheet) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index in range(1, worksheet.max_column + 1):
        values = [
            "" if cell.value is None else str(cell.value)
            for cell in list(worksheet.columns)[column_index - 1]
        ]
        max_length = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 10), 45
        )


def rebuild_workbook(input_dir: Path, output_path: Path) -> None:
    if output_path.resolve() == DEFAULT_WORKBOOK.resolve():
        raise ValueError(
            "Refusing to overwrite ICD_Database.xlsx. Use a separate output file, "
            "review it, then copy approved data manually."
        )

    sheets = load_manifest(input_dir)
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)

    for sheet in sheets:
        sheet_name = str(sheet["sheet_name"])
        csv_path = input_dir / str(sheet["csv_file"])
        if not csv_path.is_file():
            raise FileNotFoundError(f"Worksheet CSV not found: {csv_path}")

        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.sheet_state = str(sheet.get("sheet_state", "visible"))

        with csv_path.open("r", newline="", encoding="utf-8-sig") as csv_file:
            for row in csv.reader(csv_file):
                worksheet.append(row)

        format_worksheet(worksheet)
        print(f"Imported {csv_path.name} -> {sheet_name!r}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Rebuilt workbook -> {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Rebuild a formatted ICD workbook from per-sheet CSV files."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_CSV_DIR,
        help=f"CSV input directory (default: {DEFAULT_CSV_DIR})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_REBUILT,
        help=f"New workbook path (default: {DEFAULT_REBUILT.name})",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    rebuild_workbook(args.input_dir.resolve(), args.output.resolve())
